import openpyxl as ol

wb2 = ol.Workbook()
st2 = wb2.active
st2['A1'] = "电影排行榜"
second = ['名称', '评分']
st2.append(second)
st2.append(['xiao', '9.8'])
# st2.unmerge_cells('A1:B1')
wb2.save('new.xlsx')
# wb1 = ol.load_workbook('test.xlsx')
# print(wb1.sheetnames, type(wb1.sheetnames))
# wb2 = ol.Workbook()
# wb2.create_sheet('新sheet')
# wb2.save('new.xlsx')
# print(wb2.sheetnames, type(wb2.sheetnames))
# print(wb1.active)
# st = wb1['成绩表']
# print(st)
# print(st.title)
# print(st.dimensions)
# print(st.max_row)
# print(st.max_column)

# c = st.cell(row=4, column=3)
# c2 = st['c4']
# print(c2.value)
# c3 = st['A3':'C7']
# print(c3)
# print(type(c3))
# for row in c3:
#     for j in row:
#         print(j.value, end=' ')
#     print()

# for ce in st['A']:
#     print(ce.value)
# for ce in st['4']:
#     print(ce.value)
# for ce in st['A':'B']:
#     for c in ce:
#         print(c.value, end=' ')

# for ce in st['3':'5']:
#     for c in ce:
#         print(c.value, end=' ')
# st['c6'] = 100
# st.cell(row=4, cloumn=3).value = 99
# wb1.save('test.xlsx')
